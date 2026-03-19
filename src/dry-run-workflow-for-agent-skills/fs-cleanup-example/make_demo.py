from __future__ import annotations

import csv
import json
import os
import platform
import random
import textwrap
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook


# =========================
# Hard-coded project output
# =========================
ROOT_DIR = Path("example_business_project")

# Each entry:
# path, file_type, created_dt, payload
FILES = [
    {
        "path": "Analysis/2025_q1/Revenue_FINAL_v2.csv",
        "type": "csv",
        "dt": "2025-01-09 08:14:00",
        "payload": [
            ["region", "account_mgr", "jan_revenue", "feb_revenue", "mar_revenue"],
            ["North", "J. Patel", 128000, 131500, 125300],
            ["South", "lisa", 98000, 102300, 110250],
            ["west", "M TURNER", 143200, 140100, 146990],
        ],
    },
    {
        "path": "Analysis/2025_q1/rev-cleanup_NOTES.txt",
        "type": "txt",
        "dt": "2025-01-22 17:33:00",
        "payload": """
            TODO:
            - confirm West march spike
            - Lisa's numbers maybe from old extract?
            - ask finance if rebate booked twice
            - delete duplicate csv later
        """,
    },
    {
        "path": "ANALYSIS/2025_Q2/client-Margin review.xlsx",
        "type": "xlsx",
        "dt": "2025-02-11 10:05:00",
        "payload": {
            "sheet_name": "Margin",
            "rows": [
                ["Client", "Segment", "Revenue", "COGS", "MarginPct"],
                ["Acme Holdings", "Enterprise", 525000, 418000, 0.2038],
                ["BlueRiver", "SMB", 184000, 151200, 0.1783],
                ["Nova Retail", "MidMarket", 276500, 219900, 0.2047],
            ],
        },
    },
    {
        "path": "ops//weekly_stuff/ops_export_03-04-25.csv",
        "type": "csv",
        "dt": "2025-03-04 06:52:00",
        "payload": [
            ["ticket_id", "priority", "owner", "hours_open", "status"],
            [44001, "High", "Jim", 11, "Working"],
            [44002, "Low", "Tanya", 73, "Waiting"],
            [44003, "Urgent", "n/a", 4, "Escalated"],
            [44004, "Medium", "R.Khan", 26, "Closed?"],
        ],
    },
    {
        "path": "ops/weekly_stuff/READ_ME_maybe.md",
        "type": "txt",
        "dt": "2025-03-19 12:01:00",
        "payload": """
            # ops weekly stuff

            This folder has exports from the old queue pull.
            Some are clean enough to use, some are not.
            Nobody remembers which script generated them.
        """,
    },
    {
        "path": "Finance/tmp/Budget_2025_vFinal_FINAL.xlsx",
        "type": "xlsx",
        "dt": "2025-04-02 09:40:00",
        "payload": {
            "sheet_name": "Budget",
            "rows": [
                ["Dept", "Jan", "Feb", "Mar", "Apr", "Owner"],
                ["Sales", 85000, 87000, 86000, 91000, "M.B."],
                ["Ops", 43000, 42500, 44900, 45200, "Ravi"],
                ["Marketing", 39000, 40250, 39500, 41000, "Jen"],
                ["HR", 19000, 19300, 19100, 19800, "Kelly"],
            ],
        },
    },
    {
        "path": "Finance/tmp/budget_2025_USETHIS.csv",
        "type": "csv",
        "dt": "2025-04-07 15:15:00",
        "payload": [
            ["Dept", "AnnualBudget", "Comment"],
            ["Sales", 1049000, "closest to approved"],
            ["Ops", 534000, "missing one line maybe"],
            ["Marketing", 478250, "rounded"],
            ["HR", 231000, ""],
        ],
    },
    {
        "path": "client_data/ACME/ACME_contact_dump.json",
        "type": "json",
        "dt": "2025-05-06 11:28:00",
        "payload": {
            "client": "ACME",
            "contacts": [
                {"name": "S. Irving", "role": "VP Procurement", "email": "sirving@acme-example.com"},
                {"name": "Darla H", "role": "Ops Director", "email": "darla.h@acme-example.com"},
                {"name": "unknown old", "role": "?", "email": "legacy.contact@acme-example.com"},
            ],
            "source": "manual export",
        },
    },
    {
        "path": "client_data/ACME/meetingNotes_MAY.txt",
        "type": "txt",
        "dt": "2025-05-13 14:48:00",
        "payload": """
            Meeting notes:
            - client wants better reporting
            - complained about turnaround time
            - asked for revised implementation plan
            - somebody mentioned a discount but not sure who approved it
        """,
    },
    {
        "path": "client_data/bravo llc/Bravollc PIPELINE - Copy.csv",
        "type": "csv",
        "dt": "2025-06-03 13:02:00",
        "payload": [
            ["opportunity", "stage", "weighted_value", "close_est", "owner"],
            ["Renewal 2025", "Late", 120000, "2025-07-15", "Nina"],
            ["Expansion East", "Mid", 48000, "2025-08-30", "KJ"],
            ["Pilot add-on", "Early", 9000, "2025-10-12", "nobody?"],
        ],
    },
    {
        "path": "dashboards/exports/kpi_snapshot_june.xlsx",
        "type": "xlsx",
        "dt": "2025-06-26 07:57:00",
        "payload": {
            "sheet_name": "KPIs",
            "rows": [
                ["Metric", "Target", "Actual", "Status"],
                ["NPS", 45, 39, "Below"],
                ["Renewal Rate", 0.93, 0.95, "Good"],
                ["Avg Response Hrs", 6, 8.7, "Bad"],
                ["Gross Margin", 0.21, 0.205, "Close"],
            ],
        },
    },
    {
        "path": "dashboards/exports/kpiSnapshot_JUNE_old.csv",
        "type": "csv",
        "dt": "2025-06-28 18:20:00",
        "payload": [
            ["metric", "value"],
            ["NPS", 37],
            ["Renewal Rate", 0.94],
            ["Avg Response Hrs", 9.1],
            ["Gross Margin", 0.199],
        ],
    },
    {
        "path": "Sales Stuff/Q3/forecast__Q3__dirty.csv",
        "type": "csv",
        "dt": "2025-07-08 16:11:00",
        "payload": [
            ["rep", "territory", "forecast", "confidence"],
            ["Emma", "Central", 210000, "med"],
            ["Jon", "Northeast", 184500, "high"],
            ["Priya", "West", 265000, "low?"],
        ],
    },
    {
        "path": "Sales Stuff/Q3/q3_notes  final maybe.txt",
        "type": "txt",
        "dt": "2025-07-18 09:22:00",
        "payload": """
            Use Emma number from CRM not sheet.
            Priya said West forecast is sandbagged.
            Jon still missing one hospital account.
        """,
    },
    {
        "path": "HR/misc/HeadCount_2025-Aug.xlsx",
        "type": "xlsx",
        "dt": "2025-08-05 10:49:00",
        "payload": {
            "sheet_name": "HC",
            "rows": [
                ["Department", "Open Roles", "Filled Roles", "Attrition YTD"],
                ["Sales", 4, 38, 3],
                ["Ops", 2, 24, 1],
                ["Marketing", 1, 12, 2],
                ["Finance", 0, 8, 0],
            ],
        },
    },
    {
        "path": "HR/misc/people_list_really_old_but2025.csv",
        "type": "csv",
        "dt": "2025-08-14 08:03:00",
        "payload": [
            ["emp_id", "name", "dept", "status"],
            [1001, "M. Brown", "Sales", "Active"],
            [1002, "Ravi Singh", "Ops", "Active"],
            [1003, "Jen C", "Marketing", "Leave"],
            [1004, "Kelly D", "HR", "Active"],
        ],
    },
    {
        "path": "Archive_for_real/Sept/old-analysis-v3(FINAL).txt",
        "type": "txt",
        "dt": "2025-09-09 19:16:00",
        "payload": """
            Archived analysis summary:
            margin compression continued in low-value accounts.
            suspected pricing leakage in 2 segments.
            revisit after October close.
        """,
    },
    {
        "path": "Archive_for_real/Sept/raw_extract_09_09_2025.csv",
        "type": "csv",
        "dt": "2025-09-09 19:20:00",
        "payload": [
            ["id", "segment", "rev", "cost", "flag"],
            [1, "A", 12000, 10300, ""],
            [2, "B", 9800, 9100, "review"],
            [3, "C", 15400, 13200, ""],
            [4, "A", 11100, 10750, "odd"],
        ],
    },
    {
        "path": "legal-ish/contracts/contract_summary_OCT.xlsx",
        "type": "xlsx",
        "dt": "2025-10-03 11:11:00",
        "payload": {
            "sheet_name": "Contracts",
            "rows": [
                ["Client", "Term", "AutoRenew", "SpecialTerms"],
                ["ACME", "24 months", "Yes", "price floor"],
                ["Bravo LLC", "12 months", "No", "pilot carveout"],
                ["Nova Retail", "36 months", "Yes", "custom SLA"],
            ],
        },
    },
    {
        "path": "legal-ish/contracts/read-first_OR_NOT.txt",
        "type": "txt",
        "dt": "2025-10-21 17:05:00",
        "payload": """
            Not legal advice.
            Some clauses copied from prior templates.
            Final versions may be in email.
        """,
    },
    {
        "path": "YearEnd/misc numbers/YE_adj??.csv",
        "type": "csv",
        "dt": "2025-11-12 20:34:00",
        "payload": [
            ["bucket", "amount", "reason"],
            ["Revenue Deferral", -18000, "timing"],
            ["Expense Reclass", 6200, "wrong dept"],
            ["Accrual", 14300, "estimate only"],
        ],
    },
    {
        "path": "YearEnd/misc numbers/yEnd-summary-FINAL-FINAL.txt",
        "type": "txt",
        "dt": "2025-11-28 09:59:00",
        "payload": """
            Need signoff from finance.
            Numbers do not fully tie to dashboard.
            One workbook still uses pre-close data.
        """,
    },
    {
        "path": "zzz_oldbutkeep/DEC/client_rollup_2025.xlsx",
        "type": "xlsx",
        "dt": "2025-12-05 08:44:00",
        "payload": {
            "sheet_name": "Rollup",
            "rows": [
                ["Client", "2025 Revenue", "Margin", "Risk"],
                ["ACME", 1450000, 0.214, "Medium"],
                ["Bravo LLC", 544000, 0.187, "Low"],
                ["Nova Retail", 921000, 0.206, "High"],
            ],
        },
    },
    {
        "path": "zzz_oldbutkeep/DEC/DO_NOT_DELETE_maybe.json",
        "type": "json",
        "dt": "2025-12-19 22:10:00",
        "payload": {
            "status": "unknown",
            "owner": "former analyst",
            "note": "kept because it was referenced in a meeting once",
            "reviewed": False,
        },
    },
]

EXTRA_EMPTY_DIRS = [
    "random",
    "random/old",
    "temp_ish",
    "temp_ish/for-review",
    "Shared Drive Dump",
    "Shared Drive Dump/more stuff",
]


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def write_txt(path: Path, content: str) -> None:
    ensure_parent(path)
    path.write_text(textwrap.dedent(content).strip() + "\n", encoding="utf-8")


def write_csv(path: Path, rows: list[list]) -> None:
    ensure_parent(path)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def write_json(path: Path, data: dict) -> None:
    ensure_parent(path)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def write_xlsx(path: Path, payload: dict) -> None:
    ensure_parent(path)
    wb = Workbook()
    ws = wb.active
    ws.title = payload.get("sheet_name", "Sheet1")

    for row in payload.get("rows", []):
        ws.append(row)

    # Add a little extra "business mess"
    ws["H2"] = "notes"
    ws["H3"] = random.choice(
        [
            "check this later",
            "from old export",
            "seems right?",
            "copied from finance call",
            "draft only",
        ]
    )
    wb.save(path)


def set_file_times_cross_platform(path: Path, dt_str: str) -> None:
    """
    Sets modified/accessed time everywhere.
    Attempts to set creation time on Windows only.
    """
    dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
    ts = dt.timestamp()

    # Accessed + Modified
    os.utime(path, (ts, ts))

    # Creation time on Windows
    if platform.system() == "Windows":
        try:
            import ctypes
            from ctypes import wintypes

            FILE_WRITE_ATTRIBUTES = 0x0100
            OPEN_EXISTING = 3
            FILE_ATTRIBUTE_NORMAL = 0x80

            kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)

            CreateFileW = kernel32.CreateFileW
            CreateFileW.argtypes = [
                wintypes.LPCWSTR,
                wintypes.DWORD,
                wintypes.DWORD,
                wintypes.LPVOID,
                wintypes.DWORD,
                wintypes.DWORD,
                wintypes.HANDLE,
            ]
            CreateFileW.restype = wintypes.HANDLE

            SetFileTime = kernel32.SetFileTime
            SetFileTime.argtypes = [
                wintypes.HANDLE,
                ctypes.POINTER(wintypes.FILETIME),
                ctypes.POINTER(wintypes.FILETIME),
                ctypes.POINTER(wintypes.FILETIME),
            ]
            SetFileTime.restype = wintypes.BOOL

            CloseHandle = kernel32.CloseHandle
            CloseHandle.argtypes = [wintypes.HANDLE]
            CloseHandle.restype = wintypes.BOOL

            # Windows filetime = 100-ns intervals since Jan 1, 1601 UTC
            EPOCH_AS_FILETIME = 116444736000000000
            HUNDREDS_OF_NS = 10_000_000

            filetime_value = int(ts * HUNDREDS_OF_NS + EPOCH_AS_FILETIME)
            low = filetime_value & 0xFFFFFFFF
            high = filetime_value >> 32
            ft = wintypes.FILETIME(low, high)

            handle = CreateFileW(
                str(path),
                FILE_WRITE_ATTRIBUTES,
                0,
                None,
                OPEN_EXISTING,
                FILE_ATTRIBUTE_NORMAL,
                None,
            )
            if handle == wintypes.HANDLE(-1).value:
                raise OSError("CreateFileW failed")

            try:
                ok = SetFileTime(handle, ctypes.byref(ft), ctypes.byref(ft), ctypes.byref(ft))
                if not ok:
                    raise OSError("SetFileTime failed")
            finally:
                CloseHandle(handle)

        except Exception as e:
            print(f"[warn] Could not set Windows creation time for {path}: {e}")


def create_file(base_dir: Path, item: dict) -> Path:
    path = base_dir / item["path"]
    file_type = item["type"]

    if file_type == "txt":
        write_txt(path, item["payload"])
    elif file_type == "csv":
        write_csv(path, item["payload"])
    elif file_type == "json":
        write_json(path, item["payload"])
    elif file_type == "xlsx":
        write_xlsx(path, item["payload"])
    else:
        raise ValueError(f"Unsupported file type: {file_type}")

    set_file_times_cross_platform(path, item["dt"])
    return path


def create_extra_dirs(base_dir: Path) -> None:
    for d in EXTRA_EMPTY_DIRS:
        (base_dir / d).mkdir(parents=True, exist_ok=True)


def main() -> None:
    base_dir = ROOT_DIR
    base_dir.mkdir(parents=True, exist_ok=True)

    create_extra_dirs(base_dir)

    created = []
    for item in FILES:
        created_path = create_file(base_dir, item)
        created.append(created_path)

    # Add a couple of duplicate-ish junk files
    junk_files = [
        base_dir / "Analysis" / "2025_q1" / "Revenue_FINAL_v2 (1).csv",
        base_dir / "Finance" / "tmp" / "~$Budget_2025_vFinal_FINAL.xlsx",
        base_dir / "temp_ish" / "for-review" / "notes_old_old.txt",
    ]

    write_csv(
        junk_files[0],
        [
            ["region", "account_mgr", "jan_revenue"],
            ["North", "J. Patel", 128000],
            ["South", "lisa", 98000],
        ],
    )
    write_txt(junk_files[2], "old review notes\nnothing useful here\n")

    # Fake Excel lock file / temp artifact
    ensure_parent(junk_files[1])
    junk_files[1].write_text("", encoding="utf-8")

    junk_dates = [
        "2025-01-10 08:16:00",
        "2025-04-02 09:41:00",
        "2025-11-01 12:00:00",
    ]
    for p, dt in zip(junk_files, junk_dates):
        set_file_times_cross_platform(p, dt)

    print(f"Created messy project data in: {base_dir.resolve()}")
    print(f"Total files created: {len(created) + len(junk_files)}")


if __name__ == "__main__":
    main()
