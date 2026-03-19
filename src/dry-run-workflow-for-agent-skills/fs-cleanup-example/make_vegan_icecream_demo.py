from __future__ import annotations

import csv
import json
import os
import platform
import random
import textwrap
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

ROOT_DIR = Path("vegan_icecream_example_project")

FILES = [
    {
        "path": "Sales-Analysis/Q1_2025/flavor_sales_FINALish.csv",
        "type": "csv",
        "dt": "2025-01-08 08:12:00",
        "payload": [
            ["flavor", "channel", "jan_units", "feb_units", "mar_units"],
            ["Vanilla Bean Oat", "Retail", 12400, 13120, 12980],
            ["Mint Chip Coconut", "Retail", 9180, 9040, 9555],
            ["Salted Caramel Cashew", "Foodservice", 6800, 7205, 7010],
            ["Strawberry swirl", "DTC", 2450, 2620, 2815],
        ],
    },
    {
        "path": "Sales-Analysis/Q1_2025/notes_for_finance maybe.txt",
        "type": "txt",
        "dt": "2025-01-19 17:41:00",
        "payload": """
            Need to confirm promo lift on Mint Chip Coconut.
            Vanilla Bean Oat units look right but revenue may include credit memo.
            DTC refund numbers were pulled from a different export.
            leave this here for now
        """,
    },
    {
        "path": "R&D/bench tests/Flavor-Margin Review.xlsx",
        "type": "xlsx",
        "dt": "2025-02-14 09:06:00",
        "payload": {
            "sheet_name": "Margins",
            "rows": [
                [
                    "Flavor",
                    "IngredientCost",
                    "PackCost",
                    "WholesalePrice",
                    "ApproxMarginPct",
                ],
                ["Vanilla Bean Oat", 1.24, 0.38, 3.25, 0.50],
                ["Mint Chip Coconut", 1.41, 0.38, 3.45, 0.48],
                ["Salted Caramel Cashew", 1.57, 0.42, 3.85, 0.48],
                ["Lemon Poppy", 1.19, 0.37, 3.10, 0.50],
            ],
        },
    },
    {
        "path": "ops_exports/factory/Line2-output_02-27-25.csv",
        "type": "csv",
        "dt": "2025-02-27 06:49:00",
        "payload": [
            [
                "run_id",
                "sku",
                "planned_pints",
                "actual_pints",
                "downtime_min",
                "operator",
            ],
            [22011, "VBO-16", 4800, 4710, 12, "JR"],
            [22012, "MCC-16", 4300, 4218, 19, "Nadia"],
            [22013, "SCC-16", 3200, 3180, 4, "T. Lee"],
        ],
    },
    {
        "path": "ops_exports/factory/README-OLD-ish.md",
        "type": "txt",
        "dt": "2025-03-03 11:22:00",
        "payload": """
            This folder has production exports from the freezing line.
            Some files came from the old MES pull and some from manual spreadsheet entry.
            Nobody updated the naming standards.
        """,
    },
    {
        "path": "Finance/tmp/2025_Budget_vFinal_FINAL.xlsx",
        "type": "xlsx",
        "dt": "2025-03-17 14:15:00",
        "payload": {
            "sheet_name": "Budget",
            "rows": [
                ["Department", "Jan", "Feb", "Mar", "Apr", "Owner"],
                ["Sales", 78000, 79400, 80100, 81500, "M.T."],
                ["Plant Ops", 104000, 108500, 107200, 110300, "R Kumar"],
                ["Marketing", 42000, 44100, 43500, 44900, "Ari"],
                ["R&D", 25500, 26100, 25800, 26350, "Becca"],
            ],
        },
    },
    {
        "path": "Finance/tmp/budget_useThis_one.csv",
        "type": "csv",
        "dt": "2025-03-25 15:58:00",
        "payload": [
            ["Department", "AnnualBudget", "Comment"],
            ["Sales", 963000, "latest approved maybe"],
            ["Plant Ops", 1300000, "energy still estimated"],
            ["Marketing", 528500, "rounded"],
            ["R&D", 309800, ""],
        ],
    },
    {
        "path": "Retailers/WholeLeaf Market/wholeleaf contact dump.json",
        "type": "json",
        "dt": "2025-04-09 10:43:00",
        "payload": {
            "retailer": "WholeLeaf Market",
            "buyer_contacts": [
                {
                    "name": "Dana W",
                    "role": "Frozen Buyer",
                    "email": "dana.w@wholeleaf-example.com",
                },
                {
                    "name": "R. Cortez",
                    "role": "Category Manager",
                    "email": "rcortez@wholeleaf-example.com",
                },
                {
                    "name": "old entry maybe",
                    "role": "?",
                    "email": "legacy-buyer@wholeleaf-example.com",
                },
            ],
            "source": "manual crm export",
        },
    },
    {
        "path": "Retailers/WholeLeaf Market/callNotes_APRIL.txt",
        "type": "txt",
        "dt": "2025-04-18 13:16:00",
        "payload": """
            Buyer likes Salted Caramel Cashew but wants smaller case pack.
            Asked if summer seasonal can ship by first week of June.
            Mentioned freezer resets in northwest region.
            Need revised promo calendar.
        """,
    },
    {
        "path": "marketing-ish/Summer Launch/launch_budget Copy.csv",
        "type": "csv",
        "dt": "2025-05-02 09:34:00",
        "payload": [
            ["campaign", "channel", "planned_spend", "owner", "status"],
            ["Lemon Poppy launch", "Paid Social", 22000, "Ari", "active"],
            ["Store freezer clings", "Retail", 8700, "Jess", "quoted"],
            ["Influencer sampling", "Creator", 14300, "nobody?", "draft"],
        ],
    },
    {
        "path": "marketing-ish/Summer Launch/summerFlavor_brief_v3_FINAL.txt",
        "type": "txt",
        "dt": "2025-05-16 18:05:00",
        "payload": """
            Working concept:
            premium vegan ice cream with bright citrus positioning.
            avoid saying healthy too much.
            legal wants review of plant-based claim language.
        """,
    },
    {
        "path": "QA___Docs/holds/hold-report_may29.xlsx",
        "type": "xlsx",
        "dt": "2025-05-29 07:51:00",
        "payload": {
            "sheet_name": "Holds",
            "rows": [
                ["LotCode", "Flavor", "Reason", "CasesOnHold", "Owner"],
                ["L250521A", "Mint Chip Coconut", "seal variance", 118, "QA"],
                ["L250522C", "Vanilla Bean Oat", "temp excursion review", 64, "Plant"],
                ["L250525B", "Lemon Poppy", "label mismatch", 42, "Packaging"],
            ],
        },
    },
    {
        "path": "QA___Docs/holds/read_me_first or dont.txt",
        "type": "txt",
        "dt": "2025-06-01 12:11:00",
        "payload": """
            QA holds here are not always current.
            Some lots are already released but nobody moved the file.
            check email before using for leadership update.
        """,
    },
    {
        "path": "Data-Dumps/ecom/shopify_export_JUNE.csv",
        "type": "csv",
        "dt": "2025-06-12 21:09:00",
        "payload": [
            ["order_id", "sku", "qty", "gross_sales", "discounts", "ship_state"],
            [770001, "VBO-6PK", 1, 39.99, 0, "CA"],
            [770002, "MCC-6PK", 2, 79.98, 8.00, "WA"],
            [770003, "SUMMER-SAMPLER", 1, 44.99, 5.00, "OR"],
            [770004, "SCC-6PK", 1, 39.99, 0, "NY"],
        ],
    },
    {
        "path": "Data-Dumps/ecom/shopify_export_june_OLD.csv",
        "type": "csv",
        "dt": "2025-06-18 08:24:00",
        "payload": [
            ["order_id", "sku", "qty", "gross_sales"],
            [769800, "VBO-6PK", 1, 39.99],
            [769801, "MCC-6PK", 1, 39.99],
        ],
    },
    {
        "path": "Sales Stuff/Q3/forecast_Q3_dirtybutusable.csv",
        "type": "csv",
        "dt": "2025-07-07 16:27:00",
        "payload": [
            ["account", "channel", "forecast_cases", "confidence", "owner"],
            ["WholeLeaf Market", "Natural Grocery", 14200, "med", "Sam"],
            ["Fresh Basket", "Regional Grocery", 8700, "high", "Nina"],
            ["Hotel/Resort group", "Foodservice", 2600, "low?", "KJ"],
        ],
    },
    {
        "path": "Sales Stuff/Q3/aug reset notes final maybe.txt",
        "type": "txt",
        "dt": "2025-07-21 08:55:00",
        "payload": """
            WholeLeaf likely adding 2 SKUs if freezer test goes well.
            Fresh Basket wants sharper opening promo.
            Foodservice pipeline not clean. one distributor number is duplicated.
        """,
    },
    {
        "path": "People Ops/misc/Headcount_Aug2025.xlsx",
        "type": "xlsx",
        "dt": "2025-08-06 10:02:00",
        "payload": {
            "sheet_name": "HC",
            "rows": [
                ["Team", "Budgeted", "Actual", "OpenReqs"],
                ["Sales", 12, 11, 1],
                ["Plant Ops", 28, 27, 3],
                ["Marketing", 5, 5, 0],
                ["R&D", 4, 3, 1],
                ["QA", 6, 6, 0],
            ],
        },
    },
    {
        "path": "People Ops/misc/staff_list_old_but_keep.csv",
        "type": "csv",
        "dt": "2025-08-15 07:44:00",
        "payload": [
            ["emp_id", "name", "dept", "status"],
            [2001, "Sam B", "Sales", "Active"],
            [2002, "Nadia K", "Plant Ops", "Active"],
            [2003, "Ari M", "Marketing", "Active"],
            [2004, "Becca L", "R&D", "Leave"],
        ],
    },
    {
        "path": "Archive_for_real/Sept/old-flavor-analysis(FINAL).txt",
        "type": "txt",
        "dt": "2025-09-10 19:03:00",
        "payload": """
            Archived summary:
            strongest margin remains Vanilla Bean Oat.
            Salted Caramel Cashew velocity good in premium stores.
            Lemon Poppy still too early to call.
        """,
    },
    {
        "path": "Archive_for_real/Sept/raw_extract_09_10_2025.csv",
        "type": "csv",
        "dt": "2025-09-10 19:08:00",
        "payload": [
            ["sku", "week", "ship_cases", "returns_cases", "promo_flag"],
            ["VBO-16", 36, 2010, 15, ""],
            ["MCC-16", 36, 1740, 9, "promo"],
            ["SCC-16", 36, 1515, 7, ""],
            ["LEM-16", 36, 880, 3, "promo"],
        ],
    },
    {
        "path": "legal-ish/claims/claims_review_OCT.xlsx",
        "type": "xlsx",
        "dt": "2025-10-02 11:37:00",
        "payload": {
            "sheet_name": "Claims",
            "rows": [
                ["Claim", "Status", "Owner", "Notes"],
                ["Vegan", "approved", "Legal", "OK"],
                ["Dairy-Free", "approved", "Legal", "confirm supplier docs"],
                ["Plant-Based", "approved-ish", "Legal", "watch packaging context"],
                ["All Natural", "hold", "Marketing", "too broad maybe"],
            ],
        },
    },
    {
        "path": "legal-ish/claims/read-this_first_ORmaybe.txt",
        "type": "txt",
        "dt": "2025-10-23 17:12:00",
        "payload": """
            Not legal advice.
            Some copy points were pulled from previous packaging.
            Final approved wording might only be in email chain.
        """,
    },
    {
        "path": "YearEnd/misc numbers/YE_adjustments??.csv",
        "type": "csv",
        "dt": "2025-11-11 20:21:00",
        "payload": [
            ["bucket", "amount", "reason"],
            ["Trade Spend Reserve", -14500, "late retailer accrual"],
            ["Freight Reclass", 6300, "wrong cost center"],
            ["Inventory write-down", -9800, "obsolete seasonal pints"],
        ],
    },
    {
        "path": "YearEnd/misc numbers/yr-end-summary-FINAL-final.txt",
        "type": "txt",
        "dt": "2025-11-26 09:47:00",
        "payload": """
            Need CFO signoff.
            Sales deck does not fully tie to finance workbook.
            One margin sheet still uses pre-adjustment freight costs.
        """,
    },
    {
        "path": "zzz_oldbutkeep/DEC/retailer_rollup_2025.xlsx",
        "type": "xlsx",
        "dt": "2025-12-04 08:28:00",
        "payload": {
            "sheet_name": "Rollup",
            "rows": [
                ["Retailer", "2025 Net Sales", "Avg Margin", "Risk"],
                ["WholeLeaf Market", 1845000, 0.224, "Medium"],
                ["Fresh Basket", 942000, 0.208, "Low"],
                ["Green Valley Co-op", 488500, 0.231, "Low"],
                ["Hotel/Resort group", 301200, 0.183, "High"],
            ],
        },
    },
    {
        "path": "zzz_oldbutkeep/DEC/DO_NOT_DELETE_this_one.json",
        "type": "json",
        "dt": "2025-12-17 22:03:00",
        "payload": {
            "status": "unknown",
            "owner": "former ops analyst",
            "note": "kept because leadership referenced it in one meeting",
            "reviewed": False,
        },
    },
]

EXTRA_EMPTY_DIRS = [
    "random",
    "random/old",
    "temp_ish",
    "temp_ish/review-later",
    "Shared Drive Dump",
    "Shared Drive Dump/more frozen stuff",
    "innovation??",
    "innovation??/unused",
]


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def write_txt(path: Path, content: str) -> None:
    ensure_parent(path)
    path.write_text(textwrap.dedent(content).strip() + "\n", encoding="utf-8")


def write_csv(path: Path, rows: list[list]) -> None:
    ensure_parent(path)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def write_json(path: Path, data: dict) -> None:
    ensure_parent(path)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def write_xlsx(path: Path, payload: dict) -> None:
    ensure_parent(path)
    wb = Workbook()
    ws = wb.active
    ws.title = payload.get("sheet_name", "Sheet1")

    for row in payload.get("rows", []):
        ws.append(row)

    ws["H2"] = "notes"
    ws["H3"] = random.choice(
        [
            "check with finance",
            "from old export",
            "plant said this looked right?",
            "draft only",
            "copied from meeting notes",
        ]
    )

    wb.save(path)


def set_file_times_cross_platform(path: Path, dt_str: str) -> None:
    """
    Sets modified/accessed time on all platforms.
    Attempts to set creation time on Windows only.
    """
    dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
    ts = dt.timestamp()

    os.utime(path, (ts, ts))

    if platform.system() == "Windows":
        try:
            import ctypes
            from ctypes import wintypes

            FILE_WRITE_ATTRIBUTES = 0x0100
            OPEN_EXISTING = 3
            FILE_ATTRIBUTE_NORMAL = 0x80

            kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)

            CreateFileW = kernel32.CreateFileW
            CreateFileW.argtypes = [
                wintypes.LPCWSTR,
                wintypes.DWORD,
                wintypes.DWORD,
                wintypes.LPVOID,
                wintypes.DWORD,
                wintypes.DWORD,
                wintypes.HANDLE,
            ]
            CreateFileW.restype = wintypes.HANDLE

            SetFileTime = kernel32.SetFileTime
            SetFileTime.argtypes = [
                wintypes.HANDLE,
                ctypes.POINTER(wintypes.FILETIME),
                ctypes.POINTER(wintypes.FILETIME),
                ctypes.POINTER(wintypes.FILETIME),
            ]
            SetFileTime.restype = wintypes.BOOL

            CloseHandle = kernel32.CloseHandle
            CloseHandle.argtypes = [wintypes.HANDLE]
            CloseHandle.restype = wintypes.BOOL

            EPOCH_AS_FILETIME = 116444736000000000
            HUNDREDS_OF_NS = 10_000_000

            filetime_value = int(ts * HUNDREDS_OF_NS + EPOCH_AS_FILETIME)
            low = filetime_value & 0xFFFFFFFF
            high = filetime_value >> 32
            ft = wintypes.FILETIME(low, high)

            handle = CreateFileW(
                str(path),
                FILE_WRITE_ATTRIBUTES,
                0,
                None,
                OPEN_EXISTING,
                FILE_ATTRIBUTE_NORMAL,
                None,
            )
            if handle == wintypes.HANDLE(-1).value:
                raise OSError("CreateFileW failed")

            try:
                ok = SetFileTime(
                    handle, ctypes.byref(ft), ctypes.byref(ft), ctypes.byref(ft)
                )
                if not ok:
                    raise OSError("SetFileTime failed")
            finally:
                CloseHandle(handle)

        except Exception as e:
            print(f"[warn] Could not set Windows creation time for {path}: {e}")


def create_file(base_dir: Path, item: dict) -> Path:
    path = base_dir / item["path"]
    file_type = item["type"]

    if file_type == "txt":
        write_txt(path, item["payload"])
    elif file_type == "csv":
        write_csv(path, item["payload"])
    elif file_type == "json":
        write_json(path, item["payload"])
    elif file_type == "xlsx":
        write_xlsx(path, item["payload"])
    else:
        raise ValueError(f"Unsupported file type: {file_type}")

    set_file_times_cross_platform(path, item["dt"])
    return path


def create_extra_dirs(base_dir: Path) -> None:
    for d in EXTRA_EMPTY_DIRS:
        (base_dir / d).mkdir(parents=True, exist_ok=True)


def main() -> None:
    base_dir = ROOT_DIR
    base_dir.mkdir(parents=True, exist_ok=True)

    create_extra_dirs(base_dir)

    created = []
    for item in FILES:
        created_path = create_file(base_dir, item)
        created.append(created_path)

    junk_files = [
        base_dir / "Sales-Analysis" / "Q1_2025" / "flavor_sales_FINALish (1).csv",
        base_dir / "Finance" / "tmp" / "~$2025_Budget_vFinal_FINAL.xlsx",
        base_dir / "temp_ish" / "review-later" / "old_notes_old_old.txt",
        base_dir / "marketing-ish" / "Summer Launch" / "launch_budget Copy 2.csv",
    ]

    write_csv(
        junk_files[0],
        [
            ["flavor", "channel", "jan_units"],
            ["Vanilla Bean Oat", "Retail", 12400],
            ["Mint Chip Coconut", "Retail", 9180],
        ],
    )

    ensure_parent(junk_files[1])
    junk_files[1].write_text("", encoding="utf-8")

    write_txt(
        junk_files[2],
        """
            old review notes
            not sure if useful
            probably superseded by another file
        """,
    )

    write_csv(
        junk_files[3],
        [
            ["campaign", "channel", "planned_spend"],
            ["Lemon Poppy launch", "Paid Social", 22000],
            ["Store freezer clings", "Retail", 8700],
        ],
    )

    junk_dates = [
        "2025-01-09 08:14:00",
        "2025-03-17 14:16:00",
        "2025-11-03 12:00:00",
        "2025-05-03 09:40:00",
    ]

    for p, dt in zip(junk_files, junk_dates):
        set_file_times_cross_platform(p, dt)

    print(f"Created messy vegan ice cream example project in: {base_dir.resolve()}")
    print(f"Total files created: {len(created) + len(junk_files)}")


if __name__ == "__main__":
    main()
